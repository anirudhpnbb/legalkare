import os
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

# LangChain / OpenAI imports
from langchain_openai import ChatOpenAI
from langchain_core.messages import HumanMessage, SystemMessage
from langchain_core.output_parsers import StrOutputParser

# PyMuPDF (fitz) for PDFs (if needed)
try:
    import fitz
except ImportError:
    fitz = None

###############################################################################
# 1. Load environment variables for OpenAI
###############################################################################
load_dotenv()
openai_api_key = os.getenv("OPENAI_API_KEY_CLASSIFY")
model_name = os.getenv("OPENAI_MODEL_NAME")

###############################################################################
# 2. Helper to extract text from PDFs
###############################################################################
def extract_text_from_pdf(pdf_path: str) -> str:
    text_parts = []
    doc = fitz.open(pdf_path)
    for page in doc:
        text_parts.append(page.get_text())
    doc.close()
    return "\n".join(text_parts)

def read_docs_from_ik_downloads(directory_path="ik_downloads"):
    """
    Scans 'ik_downloads' folder for .txt or .pdf files.
    Yields (filename, text) for each file.
    """
    ALLOWED_EXTENSIONS = {"txt", "pdf"}

    # We'll build a list of (filename, text_content)
    docs_with_fnames = []

    for root, dirs, files in os.walk(directory_path):
        for file in files:
            file_ext = file.rsplit(".", 1)[-1].lower()
            if file_ext in ALLOWED_EXTENSIONS:
                file_path = os.path.join(root, file)
                if file_ext == "txt":
                    with open(file_path, "r", encoding="utf-8") as f:
                        text_content = f.read()
                    docs_with_fnames.append((file, text_content))
                elif file_ext == "pdf":
                    if not fitz:
                        print(f"PyMuPDF not installed; skipping PDF '{file}'.")
                        continue
                    text_content = extract_text_from_pdf(file_path)
                    docs_with_fnames.append((file, text_content))

    return docs_with_fnames

###############################################################################
# 4. LLM function to determine case type + winner in one shot
###############################################################################
def llm_classify_case_and_winner(documents_list, model_name):
    """
    Takes a list of doc texts, uses a ChatOpenAI model to:
      - Classify the case as 'criminal', 'civil', or 'unknown'.
      - Determine who won:
        * if criminal => 'state', 'defense', or 'unknown'
        * if civil => 'plaintiff', 'defendant', or 'unknown'
        * if unknown => 'unknown'
    Returns a single string like: "case_type: criminal, winner: defense"
    """

    # Instantiate the ChatOpenAI model
    model_llm = ChatOpenAI(model=model_name, openai_api_key=openai_api_key)

    # Prepare a parser for string output
    parser = StrOutputParser()

    # Combine all text into a single context
    context = " ".join(documents_list)

    # System instructions for the AI
    system_instructions = (
        "You are a legal expert AI. You will be given text from a legal case. "
        "Your job is to classify whether the case is 'criminal', 'civil', or 'unknown', "
        "and then determine who won. If it is a criminal case, the winner can only be 'state', "
        "'defense', or 'unknown'. If it is a civil case, the winner can be 'plaintiff', "
        "'defendant', or 'unknown'. If the case type is unknown, the winner is also 'unknown'. "
        "Return your answer in the format: case_type: X, winner: Y"
    )

    # User message
    user_prompt = (
        f"Case text:\n{context}\n\n"
        "Please identify the case type (criminal, civil, or unknown) and the winner "
        "(in one word). Return the result exactly as:\ncase_type: <type>, winner: <side>\n"
    )

    messages = [
        SystemMessage(content=system_instructions),
        HumanMessage(content=user_prompt)
    ]

    # Use a pipeline: LLM -> parser
    chain = model_llm | parser
    result = chain.invoke(messages)

    return result



def store_result_in_excel(file_name, classification_result, excel_path="output.xlsx"):
    """
    Appends a row to 'excel_path' with [file_name, classification_result].
    Creates a new workbook if 'excel_path' doesn't exist.
    """
    if os.path.isfile(excel_path):
        workbook = load_workbook(excel_path)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        # Optionally add headers:
        if sheet.max_row == 1 and not sheet.cell(row=1, column=1).value:
            sheet["A1"] = "Filename"
            sheet["B1"] = "Classification"

    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1).value = file_name
    sheet.cell(row=next_row, column=2).value = classification_result

    workbook.save(excel_path)
    print(f"Stored result for '{file_name}' => '{classification_result}' in '{excel_path}'")



###############################################################################
# 5. Example 'main' usage: read docs from 'ik_downloads', classify each
###############################################################################
if __name__ == "__main__":

    docs_with_fnames = read_docs_from_ik_downloads("ik_downloads")
    if not docs_with_fnames:
        print("No docs found in 'ik_downloads'.")
        exit(0)

    for (file_name, text) in docs_with_fnames:
        # Pass the single doc text as a list to the LLM function
        try:
            result = llm_classify_case_and_winner([text], model_name)
        except:
            result = "Undetermined"

        print(f"File: {file_name} => Classification: {result}")

        # Store in Excel
        store_result_in_excel(file_name, result, excel_path="my_results.xlsx")